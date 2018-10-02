from tkinter import *
from datetime import *
from tkinter import messagebox, DISABLED
from sendmails import OtprEmail
import openpyxl
import sqlite3




class Helper:
    filename = 0
    mainFrame = Tk()   
    VidProd = Entry() 
    Dogovor = Entry()
    prices = Entry()
    Zalob = Entry()
    DateText = Label()
    SmenaText = Label()
    Zarab = Label()
    Procent = Label()
    VidProdLabel = Label()
    DogovorLabel = Label()
    PriceLabel = Label()
    DogovorButton = Button()
    ZalobLabel = Label()
    ZalobButton = Button()
    EmailButton = Button()
    SmenButton = Button()
    AuthButton = Button()
    login = Entry()
    passw = Entry()
    loginLabel = Label()
    passwLabel = Label()
  
    def authlog(self,pas,log):
        connects = sqlite3.connect("db.db")
        cursor_log = connects.cursor()
        cursor_pas = connects.cursor()
        sql = "SELECT username FROM users"
        cursor_log.execute(sql)
        sql = "SELECT pass FROM users"
        cursor_pas.execute(sql)
        for find in cursor_log:
            if((log,)==find):
                for find in cursor_pas:
                    if((pas,)==find):
                        self.hideElem(1)
                        self.mainFrame.geometry("600x350+500+300")
                        self.CreateElem(self.mainFrame,'calibri 12',DISABLED)

    
    def __init__(self,filename):
            self.filename=filename
            self.CreateElem(self.mainFrame,'calibri 12',DISABLED)
            self.hideElem(0)
            self.Authorization()
            self.mainFrame.resizable(False,False)
            self.mainFrame.mainloop()
            
            
    #Скрытие элментов, 0 - элементы основной формы, 1- элементы авторизации       
    def hideElem(self,types):
        if types==0:
            self.VidProd.place_forget()
            self.Dogovor.place_forget()
            self.prices.place_forget()
            self.Zalob.place_forget()
            self.DateText.place_forget()
            self.SmenaText.place_forget()
            self.Zarab.place_forget()
            self.Procent.place_forget()
            self.VidProdLabel.place_forget()
            self.DogovorLabel.place_forget()
            self.PriceLabel.place_forget()
            self.DogovorButton.place_forget()
            self.ZalobLabel.place_forget()
            self.ZalobButton.place_forget()
            self.EmailButton.place_forget()
            self.SmenButton.place_forget()
        elif types==1:
            self.AuthButton.place_forget()
            self.login.place_forget()
            self.passw.place_forget()
            self.loginLabel.place_forget()
            self.passwLabel.place_forget()



    #Авторизация
    def Authorization(self):
        def AuthCommand():
            self.authlog(self.passw.get(),self.login.get())
        self.mainFrame.geometry("250x140+600+300")
        self.AuthButton = Button(text="Войти",width=30,command=AuthCommand)
        self.AuthButton.place(x=15,y=105)
        self.login = Entry(width=36)
        self.login.place(x=15,y=35)
        self.passw = Entry(width=36)
        self.passw.place(x=15,y=75)
        #Надписи
        loginLabel = Label(text="Введите логин")
        loginLabel.place(x=15,y=10)
        passwLabel = Label(text="Введите пароль")
        passwLabel.place(x=15,y=54)
    
    #Выгрузка таблиц
    def ReadExcel(self,variant):
        wbb = openpyxl.load_workbook(filename=self.filename,data_only=True)
        sheets = wbb['Лист1']
        if variant==0:
            smena = sheets['AG2'].value
            return smena
        elif variant==1:
            smena = sheets['AI2'].value
            return smena
        elif variant==2:
            smena = sheets['AK2'].value
            return smena
    
    

    def RunEmail(self):
        OtprEmail('Ночной отчет','kostyruzavin@mail.ru','avgkoster@gmail.com',self.filename,
                  'smtp.mail.ru','kostyruzavin@mail.ru','34650qqQQqwerty')
     #Добавление договора
    def NewDogovor(self):
        i=1
        if (len(self.VidProd.get())!=0 and len(self.Dogovor.get())!=0 and len(self.prices.get())!=0):
            wb = openpyxl.load_workbook(filename=self.filename,data_only=False)
            sheet = wb['Лист3']
            while(sheet['B'+str(i)].value!=None):
                i+=1
            sheet['A'+str(i)].value = datetime.now().strftime("%d.%m.%y")
            sheet['B'+str(i)].value = 'Рузавин К.В '
            sheet['C'+str(i)].value = self.VidProd.get()
            sheet['D'+str(i)].value = self.Dogovor.get()
            sheet['E'+str(i)].value = self.prices.get()
            sheet['F'+str(i)].value = float(self.prices.get())*0.04
            sheet['G'+str(i)].value = sheet['E'+str(i)].value
            wb.save(self.filename)
            self.CreateElem(self.mainFrame)
        else:
            messagebox.showerror("Ошибка", "Пожалуйста заполните все поля")



    #Счет процентов
    def FindProc(self):
        sum=0;i=1
        wb = openpyxl.load_workbook(filename=self.filename,data_only=True)
        sheet = wb['Лист3']
        while(sheet['B'+str(i)].value!=None):
            if sheet['B'+str(i)].value=='Рузавин К.В ':
                sum+=sheet['F'+str(i)].value
            i+=1
        return int(sum)



    #Элементы
    def CreateElem(self,Frame=mainFrame,typefont="Verdana 12",state=NORMAL):
        self.VidProd = Entry(width=30)
        self.VidProd.place(x=15,y=110)
        self.Dogovor = Entry(width=25)
        self.Dogovor.place(x=220,y=110)
        self.prices = Entry(width=10)
        self.prices.place(x=395,y=110)
        self.Zalob = Text(width=80,height=5)
        self.Zalob.place(x=15,y=160)
        #Новая смена
        self.SmenButton = Button(Frame,width=70,text='Новая смена',
                        font=typefont,command=self.NewSmena)
        self.SmenButton.place(x=15,y=45)
        #Сегодняшняя дата
        self.DateText = Label(Frame,font=typefont,
                         text="Сегодня: "+
                         str(datetime.now().strftime("%d.%m.%y"))+" |")
        self.DateText.place(x=10,y=10)
        #Сегодняшняя смена
        self.SmenaText = Label(Frame,font=typefont,
                         text="Смена: "+
                         str(self.ReadExcel(0))+"  |")
        self.SmenaText.place(x=145,y=10)
        #Текущий заработок
        self.Zarab = Label(Frame,font=typefont,
                      text="Текущий заработок: "+str(self.ReadExcel(1))+"руб. |")
        self.Zarab.place(x=235,y=10)
        self.Procent = Label(Frame, font=typefont,
                      text="Процент: " + str(self.FindProc())+"руб.")
        self.Procent.place(x=460, y=10)
        #Ввод нового договора   
        self.VidProdLabel = Label(Frame, font=typefont,
                      text="Вид продажи: ")
        self.VidProdLabel.place(x=15, y=85)
        self.DogovorLabel = Label(Frame, font=typefont,
                      text="Номер договора: ")
        self.DogovorLabel.place(x=225, y=85)
        self.PriceLabel = Label(Frame,font=typefont,
                           text='Цена: ')
        self.PriceLabel.place(x=405,y=85)
        self.DogovorButton = Button(Frame,width=11,height=-5,text='Добавить',
                            font='calibri 12',command=self.NewDogovor,state=state)
        self.DogovorButton.place(x=485,y=95)
        #Очистка договоров
        self.VidProd.delete(0,END)
        self.Dogovor.delete(0,END)
        self.prices.delete(0,END)
        self.VidProd.state=state
        self.Dogovor.state=state
        self.prices.state=state
        #Добавление жалоб
        self.ZalobLabel = Label(Frame,font = typefont,
                           text="Добавление замечаний: ")
        self.ZalobLabel.place(x=15,y=130)
        self.ZalobButton = Button(text="Добавить замечание(Новых замечаний нет)",width=93,
                             command=self.NewZalob,state=state)
        self.ZalobButton.place(x=15,y=255)
        #Отправка почты
        self.EmailButton = Button(state=state,text="Отправить отчет",width=93,command=self.RunEmail)
        self.EmailButton.place(x=15,y=295)

        #Загрузка таблиц
    def NewSmena(self):
        bools = 0
        wb = openpyxl.load_workbook(filename=self.filename,data_only=True)   
        sheet = wb['Лист1']
        for get in sheet["A2:AE2"]:
            if bools==1:
                break
            for i in range(1,30):
                if get[i].value=='x':
                    sheet.cell(row=2, column=i+1).value = 1
                    sheet["AG2"].value=sheet["AG2"].value+1
                    sheet["AI2"].value=sheet["AG2"].value*sheet["AH2"].value
                    wb.save('fick1.xlsx')
                    self.CreateElem(self.mainFrame,'calibri 12',NORMAL)
                    self.SmenButton['state']=DISABLED
                    bools=1
                    break

    #Добавление жалоб
    def NewZalob(self):
        i=1
        wb = openpyxl.load_workbook(filename=self.filename,data_only=False)
        sheet = wb['Лист2']
        while(sheet['B'+str(i)].value!=None):
            i+=1
        if(len(self.Zalob.get("1.0",'end-1c'))!=0):   
            sheet['A'+str(i)].value = datetime.now().strftime("%d.%m.%y")
            sheet['B'+str(i)].value = "Рузавин К.В"
            sheet['C'+str(i)].value = self.Zalob.get("1.0",'end-1c')
        else:
            sheet['A'+str(i)].value = datetime.now().strftime("%d.%m.%y")
            sheet['B'+str(i)].value = "Рузавин К.В"
            sheet['C'+str(i)].value = "Новых замечаний нет"    
        wb.save(self.filename)
        self.CreateElem(self.mainFrame)


           
if __name__=='__main__':
    newWindow = Helper('fick1.xlsx')
  
    

    

    


